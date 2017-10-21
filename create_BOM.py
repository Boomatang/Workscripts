from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from pathlib import Path
import sys
from openpyxl.utils.exceptions import InvalidFileException
import datetime

from pony.orm import *

FIRST_DATA_ROW = 2
ITEM_NUMBER = "A"
DESCRIPTION = "D"
LENGTH = "H"
QTY = "I"
BOUNDING_LENGTH = "E"
BOUNDING_WIDTH = "F"
BOUNDING_THICKNESS = "G"
SPECIALS = ['GRP',]
SHEET_AREA = {'plate' : (2500*1250),
              'GRP': (1010*4000)}

BOM_template = "Boom Boom version 996323"

TESTING = True 
TEST_FILE = "/home/boomatang/Projects/WorkScripts/Temp/Broken_test_flie.xlsx"
# Some database set up stuff
db = Database()

class Section(db.Entity):
    id = PrimaryKey(int, auto=True)
    name = Required(str, unique=True)
    sizes = Set('Section_Sizes')

class Section_Sizes(db.Entity):
    id = PrimaryKey(int, auto=True)
    section = Required(Section)
    length = Required(int)

def create_db():
    db.bind("sqlite", "data.sqlite", create_db=True)
    db.generate_mapping(create_tables=True)

# other stuff

def trim_name(name):
    if name.startswith('"') and name.endswith('"'):
        return name[1:-1]
    else:
        return name

def get_time_value():
    return datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")

def file_name(file_path):

    parent = file_path.parent
    
    file_ext = ".xlsx"
    time_value = get_time_value()

    if TESTING:
        print(file_path.parent)
        bom = "Testing BOM "
    else:
        bom = "FS BOM "

    new_file_name = Path(parent, bom + time_value + file_ext)

    return str(new_file_name)

def get_beam_length(section):
    output = database_section_size_look_up(section)
    
    while len(output) == 0:
        print("No sections sizes where found for {}.".format(section))
        insert_section_sizes(section)
    
        output = database_section_size_look_up(section)

    # TODO this should be filled out better to give back all the sizes but for now its ouly going to give back the firsts in the list.
    return output[0].length

@db_session
def insert_section_sizes(section):
    print("""
    Please read the following.
        You will be able to enter as many sizes as you wish.
        To exit the entering just give a blank value.
        Only numbers that are base ten will be accepted.
        Numbers that are floats will be NOT converted to intagers.
    """)

    run = True
    db_section = None

    while run:
        length = input("Beam Length >> ")

        if len(length) == 0:
            print("Blank entry")
            run = False

        else:
            if is_int(length):
                if db_section is None:
                    db_section = Section(name=section)
                    db_section_size = Section_Sizes(section=db_section, length=int(length))
                else:
                    db_section_size = Section_Sizes(section=db_section, length=int(length))

            else:
                print("{} is not a valid number".format(length))

def is_int(number):
    try:
       number =  int(number)
       return True
    except ValueError:
        return False

@db_session
def database_section_size_look_up(section):
    """
    Looking up the leghts for a give section size
    retun list[int]
    """

    print("Section sizes for {} are {}".format(section, select(s.length for s in Section_Sizes if s.section.name == section)[:]))
    return select(s for s in Section_Sizes if s.section.name == section)[:]

def get_file_path():
    print("\n*** Please only use excel files created with the template {}. ***\n".format(BOM_template))
    output = input("Full path to file: ")
    if TESTING:
        print('In test mode, user input ingored')
        output = TEST_FILE
        print(TEST_FILE)
    output = Path(trim_name(output))
    if not output.is_file():
        print("Please start again and enter a real file path!")
        sys.exit(0)
    
    return output

def load_file(format_file: Path):

    try:
        wb = load_workbook(str(format_file))
        print("Working on {}.".format(format_file.name))
   
    except InvalidFileException as err:
        
        print(err)
        print("Please use the correct file format.")

    return wb

def get_parent_qty(page):
    start_row = FIRST_DATA_ROW
    output = []

    while start_row <= page.max_row:
        value = (page.cell(row=start_row, column=column_index_from_string(ITEM_NUMBER)).value, page.cell(row=start_row, column=column_index_from_string(QTY)).value)
        output.append(value)
        start_row += 1

    return output

def get_description_items(page):
    output = {}

    for row in page:
        value = {
                'item_number': row[column_index_from_string(ITEM_NUMBER) - 1].value,
                'description': row[column_index_from_string(DESCRIPTION) - 1].value,
                'bounding_length': row[column_index_from_string(BOUNDING_LENGTH) - 1].value,
                'bounding_width': row[column_index_from_string(BOUNDING_WIDTH) - 1].value,
                'bounding_thickness': row[column_index_from_string(BOUNDING_THICKNESS) - 1].value,
                'length': row[column_index_from_string(LENGTH) - 1].value,
                'qty': row[column_index_from_string(QTY) - 1].value
                }

        if value['description'] in output.keys():
            output[value['description']].append(value)
        else:
            if value['description'] is not None:
                output.setdefault(value['description'], [value])

    return output

def get_data(workbook):
    sheet = workbook[workbook.get_sheet_names()[0]]
    
    parent_qty = get_parent_qty(sheet)
    items = get_description_items(sheet)
    
    return parent_qty, items

def get_parents(unit):
        more_family = True
        parents = []
        parent = str(unit['item_number']).split('.')
        parent = '.'.join(parent[:-1])
        while more_family:
            parents.append(parent)
            if len(parent) <= 1:
                more_family = False
            parent = parent.split('.')
            parent = '.'.join(parent[:-1])

        return parents

def update_unit_qty(unit, unit_parents, world):
    multiplyer = 1
    for parent in unit_parents:
        for person in world:
            if parent == str(person[0]):
                multiplyer = multiplyer * int(person[1])

    unit['qty'] = unit['qty'] * multiplyer

def update_item_qty(world, items):
    
    for key in items.keys():
        for unit in items[key]:
            unit_parents = get_parents(unit)
            update_unit_qty(unit, unit_parents, world)

def format_input_data(format_file: Path):
    '''
    Formats the input file data to give actually numbers of units required
    '''
    wb = load_file(format_file)
    parents, items = get_data(wb)
    print("Closing template file")
    wb.close()

    print("Formatting data..")
    update_item_qty(parents, items) 
    
    return items

def find_special(bom):
    print('Looking for special types as listed')
    output = {}
    for value in SPECIALS:
        for key in bom.keys():
            if key.startswith(value):
                output[key] = bom[key]
    
    for key in output.keys():
        bom.pop(key)
    
    return output

def find_plates(bom):
    print('Looking for plates')
    output = {}
    
    for key in bom.keys():
        if bom[key][0]['length'] is None:
            output[key] = bom[key]

    for key in output.keys():
        bom.pop(key)
    
    return output

def find_sections(bom):
    print('Looking for sections')
    output = {}
    
    for key in bom.keys():
        
        if bom[key][0]['length'] is not None:
            try:
                float(bom[key][0]['length'])
                output[key] = bom[key]
            except ValueError as err:
                print()
                print("!!!! Fault Error !!!!")
                print(err)
                print("!!!!!!!!!!!!!!!!!!!!!")
                print()
        #else:
            #print(bom[key])
    for key in output.keys():
        bom.pop(key)
    
    return output

def sort_BOM(bom):
    special = find_special(bom)
    plates = find_plates(bom)
    sections = find_sections(bom)
    print('The following types could not be classified : {}'.format([i for i in bom.keys()]))

    return special, plates, sections

def refine_plates(plates):
    output = {}
    to_remove = []

    for plate in plates.keys():
        try:
            output[plate] = {
                    'description': plates[plate][0]['description'],
                    'total area':  plates[plate][0]['bounding_length']*plates[plate][0]['bounding_width'],
                    'size': '{} x {}'.format(plates[plate][0]['bounding_length'],plates[plate][0]['bounding_width']),
                    'thickness': plates[plate][0]['bounding_thickness'],
                    'qty': 0
                        }
            for unit in plates[plate]:
                output[plate]['qty'] += unit['qty']
        except TypeError:
            print("Plate error found with {}.".format(plate))
    
    return output

def plate_area_by_thickness(plates):
    output = {}

    for value in plates.values():
        if value['thickness'] in output.keys():
            output[value['thickness']] = output[value['thickness']] + (value['total area'] * value['qty'])
        else:
            output.setdefault(value['thickness'], (value['total area'] * value['qty']))
    return output

def number_of_sheets(area_list, sheet_area):

    for value in area_list.keys():

       area_list[value] = round(area_list[value] / sheet_area, 3)

def work_with_plates(plates):
    plates = refine_plates(plates)
    plate_area = plate_area_by_thickness(plates)
    number_of_sheets(plate_area, SHEET_AREA['plate'])
    return (plates, plate_area)

def percentage(whole, part):
    try:
        return round(100 * float(part)/float(whole), 3)
    except :
        print("There is an error")
        return 0

def flip_percentage(percentage_value):
    return round(100 - percentage_value, 3)

def remove_possible_lenght_errors(sections):
    output = []
    for section in sections:
        if section['length'] is not None:
            output.append(section)
    return output

def work_out_beam_cuts(sections, length):
    beams = []
    checked = 0
    counter = 1000
    sections = remove_possible_lenght_errors(sections)
    def item_length(x):
        return float(x['length'])


    sections.sort(key=item_length)
    sections.reverse()

    total_units = 0

    for item in sections:
        total_units += item['qty']

    while total_units > 0:
        beam = length
        beam_data = []
        for item in sections:
            while (beam - float(item['length'])) >= 0 and item['qty'] > 0:
                beam_data.append((item['item_number'], round(float(item['length']), 2)))
                beam = beam - float(item['length'])
                item['qty'] = item['qty'] - 1
                #print('{} has {} qty left'.format(item['item_number'], item['qty']))
                total_units -= 1
                

        value = flip_percentage(percentage(length, round(beam)))
        beams.append((beam_data, (length, round(beam), value)))
        counter -= 1

        if counter == 0:
            # beam is to short for the parts
            print("################\nBeams are too short for the parts\n#################")
            break
    return beams

def work_with_sections(sections):
    output = {}

    for key in sections.keys():
        size = sections[key]
        beam_length = get_beam_length(key)
        beams = work_out_beam_cuts(size, beam_length)
        output.setdefault(key, beams)
    return output

class BOM_Workbook():
    def __init__(self, filename, base_file, materials):
        self.filename = filename
        self.base_file = base_file

        self.beams = materials[0]
        self.plate = materials[1]
        self.special = materials[2]

        self.material_heading_row = 6
        self.overview_data_first_column = 2
        self.start_row = None
        
        self.wb = Workbook()
        self.cover = self.set_cover_page()

    def create(self):
        self.create_cover_page()
        print("Cover page created.")
        self.add_material_type_pages()
        self.save_file()

    def save_file(self):
        self.wb.save(filename=self.filename)
        print("File saved.")
        self.wb.close()
        print("File closed.")

    def add_material_type_pages(self):
        if self.beams is not None:
            self.add_beam_type_pages()
            print('Beam type pages added')

    def add_beam_type_pages(self):
        for beam in self.beams:
            self.create_beam_page(beam)

    def add_beam_title(self, sheet, title):
        sheet.cell(row=1, column=1).value = title

    def set_beam_overview_data(self, sheet, row, data):
        sheet.cell(row=row, column= 1).value = "Lenght : {}".format(data[0])
        sheet.cell(row=row, column= 2).value = "Waste : {}".format(data[1])
        row += 1
        sheet.cell(row=row, column= 1).value = "Percentage : {}%".format(round(data[2], 1))

    def set_beam_detail_data(self, sheet, row, data):
        column = 3
        
        for beam in data:
            self.set_beam_cell_values(sheet, row, column, beam)
            column += 1

    def set_beam_cell_values(self, sheet, row, column, beam):
        sheet.cell(row=row, column=column).value = beam[0]
        sheet.cell(row=row + 1, column=column).value = beam[1]

    def create_beam_page(self, title):
        ws = self.wb.create_sheet(title=title)
        self.add_beam_title(ws, title)
        start_row = 3
        for beam in self.beams[title]:
            self.set_beam_overview_data(ws, start_row, beam[1])
            self.set_beam_detail_data(ws, start_row, beam[0])
            start_row += 3

    def create_cover_page(self):
        self.set_created_date()
        self.set_base_file()
        self.enter_materials_used()
        self.enter_plates_used()

    def enter_plates_used(self):
        self.start_row = self.max_number_of_rows() + 2
        self.add_plate_header_information()

    def add_plate_header_information(self):
        self.add_plate_warning_massage()
        headings = ['Plate Thickness', 'No. of Sheets (8x4)']
        self.set_material_headings(self.start_row, self.overview_data_first_column, headings)
        self.start_row += 1
        self.show_plate_used()

    def show_plate_used(self):
        plates = self.format_overview_plate_data()

        def plate_sort(x):
            return x[0]

        plates.sort(key=plate_sort)

        for plate in plates:
            self.add_plate_overview_entry(plate)

    def add_plate_overview_entry(self, plate):
        column = self.overview_data_first_column
        for item in plate:
            cell = self.cover.cell(row=self.start_row, column=column)
            if column == self.overview_data_first_column:
                cell.value = str(item) + 'mm'
            else:
                cell.value = item
            column += 1
        self.start_row += 1

    def format_overview_plate_data(self):
        output = []

        for item in self.plate[1].items():
            output.append(item)

        return output

    def add_plate_warning_massage(self):
        massage = "The plate usage is taken from the bounding box of the part in question. This means that the nested usage maybe much lower. Also this is worked outed by using the area of the bounding boxes, no allowaces is give for part sizes. For these reason this should be used as a guide only."

        self.cover.cell(column=self.overview_data_first_column, row=self.start_row).value = massage
        self.start_row += 2


    def max_number_of_rows(self):
        return self.cover.max_row


    def set_base_file(self):
        self.cover['A1'] = "Base File"
        self.cover['B1'] = str(self.base_file)

    def set_created_date(self):
        create_time = datetime.datetime.today().strftime("%H:%M %d %h %y") 
        self.cover['A3'] = "Date Created"
        self.cover['B3'] = create_time

    def set_cover_page(self):
        worksheet = self.wb.active
        worksheet.title = "Cover Page"
        return worksheet

    def enter_materials_used(self):
        heading_row = self.material_heading_row
        start_column = self.overview_data_first_column
        self.set_material_headings(heading_row, start_column)
        self.show_material_overview(heading_row + 1, start_column)

    def set_material_headings(self, row, col, headings=None):
        if headings is None:
            headings = ["Materials", "Length", "QTY", "Usage Percentage"]
        start_column = col
        for heading in headings:
            self.cover.cell(row=row, column=start_column).value = heading 
            start_column += 1

    def get_average_percentage(self, item_list):
        counter = 0
        total = 0
        for items in item_list:
            total += items[1][2]
            counter += 1

        return round(total / counter, 2)

    def format_beam_overview_data(self):
        output = []
        for key in self.beams.keys():
            name = key
            percentage = self.get_average_percentage(self.beams[key])
            length = self.beams[key][0][1][0]
            count = len(self.beams[key])
            output.append((name, length, count, percentage))

        return output

    def format_overview_cell_data(self, row, column, value):
        self.cover.cell(row=row, column=column).value = value

    def show_material_overview(self, row, col):
        start_row = row
        beams = self.format_beam_overview_data()

        def beam_sort(x):
            return x[0]

        beams.sort(key=beam_sort)
        beams.reverse()

        for beam in beams:
            start_column = col
            for value in beam:
                self.format_overview_cell_data(start_row, start_column, value)
                start_column += 1
            start_row += 1

def run():
    if TESTING:
        print("In tset mode.. Using file : \n{}".format(TEST_FILE))
    file_path = get_file_path()
    BOM = format_input_data(file_path)
    special, plates, sections = sort_BOM(BOM)
    
    # there needs to be more work done to the special stuff
    special_data = special

    plate_data = work_with_plates(plates)

    print('Plate data collected..')

    section_data = work_with_sections(sections)
    materials = (section_data, plate_data, special_data)
    print('Section data collected..')
    bom = BOM_Workbook(file_name(file_path), file_path, materials)
    bom.create()

if __name__ == "__main__":
    create_db()
    run()
